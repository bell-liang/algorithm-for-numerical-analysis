# -*- coding: utf-8 -*-
"""
Created on Thu Nov 21 22:38:49 2019

@author: bell liang
"""

import openpyxl
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('跳变点电流.xlsx')
sheetnames = wb.sheetnames

sheets = []
for sheet in sheetnames:
    sheets.append(wb[sheet])

print(sheets)

contents = []

for sheet in sheets:
    columns = sheet.columns
    content = []
    for aol in columns:
        a = [x.value for x in aol]
        content.append(a)
    contents.append(content)

xs = []
ys = []

for i in range(10):
    y = []
    y.append(contents[i][0][2:])
    y.append(contents[i][2][2:])
    ys.append(y)

print(ys)

for i in range(10):
    xs.append(contents[i][1][2:])

print(xs)

plt.figure()
for i in range(4):
    plt.subplot(2, 2, i+1)
    plt.plot(xs[i], ys[i][0], label='X+')
    plt.plot(xs[i], ys[i][1], label='Y+')
    plt.legend()

plt.figure()
for i in range(4, 8):
    plt.subplot(2, 2, i-4+1)
    plt.plot(xs[i], ys[i][0], label='X+')
    plt.plot(xs[i], ys[i][1], label='Y+')
    plt.legend()
plt.figure()
for i in range(8, 10):
    plt.subplot(2, 2, i-8+1)
    plt.plot(xs[i], ys[i][0], label='X+')
    plt.plot(xs[i], ys[i][1], label='Y+')
    plt.legend()
    
plt.show()